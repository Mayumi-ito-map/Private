"""
kml_to_index_matched_20241009.geojson などの FeatureCollection を、
geojson/国分類.xlsx の地域別コード一覧に基づき 5 地域 + 未分類へ分割する。

- 国分類.xlsx の列 cn000_asia, cn100_europe, cn200_africa, cn300_america, cn450_oceania
  の 2 行目以降に並ぶコード（ISO alpha-3 等）が、properties["country"] に部分一致すれば
  その地域のファイルへ。
- いずれの列のコードにも一致しない場合は 500_no_country.geojson へ。
- 複数地域にマッチしうる場合は、アジア→欧州→アフリカ→アメリカ→オセアニアの順で最初の 1 つに割り当て。
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

REGION_ORDER = [
    "cn000_asia",
    "cn100_europe",
    "cn200_africa",
    "cn300_america",
    "cn450_oceania",
]

NO_COUNTRY_FILE = "500_no_country"


def project_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def load_region_codes(xlsx_path: Path) -> dict[str, list[str]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h).strip() if h is not None else "" for h in header_row]
        col_to_idx: dict[str, int] = {}
        for i, name in enumerate(header):
            if name in REGION_ORDER:
                col_to_idx[name] = i
        missing = [c for c in REGION_ORDER if c not in col_to_idx]
        if missing:
            raise ValueError(
                f"国分類.xlsx に次の列がありません: {missing}. 実際の列: {header}"
            )
        result: dict[str, list[str]] = {r: [] for r in REGION_ORDER}
        seen: dict[str, set[str]] = {r: set() for r in REGION_ORDER}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            for col in REGION_ORDER:
                idx = col_to_idx[col]
                if idx >= len(row):
                    continue
                v = row[idx]
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                u = s.upper()
                if u not in seen[col]:
                    seen[col].add(u)
                    result[col].append(u)
    finally:
        wb.close()
    return result


def classify_country(country_raw: object, region_codes: dict[str, list[str]]) -> str | None:
    """一致した REGION_ORDER のキーを 1 つ返す。未一致は None。"""
    if country_raw is None:
        return None
    s = str(country_raw).strip()
    if not s:
        return None
    hay = s.upper()
    for region in REGION_ORDER:
        for code in region_codes[region]:
            if code in hay:
                return region
    return None


def split_features(
    data: dict,
    region_codes: dict[str, list[str]],
) -> dict[str, list[dict]]:
    if data.get("type") != "FeatureCollection":
        raise ValueError('入力は type が "FeatureCollection" の GeoJSON を想定しています。')
    features = data.get("features") or []
    buckets: dict[str, list[dict]] = {r: [] for r in REGION_ORDER}
    buckets[NO_COUNTRY_FILE] = []

    for feat in features:
        props = feat.get("properties") if isinstance(feat, dict) else None
        props = props if isinstance(props, dict) else {}
        region = classify_country(props.get("country"), region_codes)
        if region is None:
            buckets[NO_COUNTRY_FILE].append(feat)
        else:
            buckets[region].append(feat)

    return buckets


def write_collection(features: list[dict], path: Path) -> None:
    out = {"type": "FeatureCollection", "features": features}
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main() -> None:
    root = project_root()
    parser = argparse.ArgumentParser(
        description="GeoJSON を国分類.xlsx の 5 地域 + 未分類に分割する。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=root / "geojson" / "kml_to_index_matched_20241009.geojson",
        help="入力 GeoJSON（FeatureCollection）",
    )
    parser.add_argument(
        "--classification-xlsx",
        type=Path,
        default=root / "geojson" / "国分類.xlsx",
        help="地域別コード一覧（1 行目が列名、2 行目以降がコード）",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=root / "geojson" / "by_region",
        help="分割ファイルの出力先ディレクトリ",
    )
    args = parser.parse_args()
    input_path: Path = args.input.expanduser().resolve()
    xlsx_path: Path = args.classification_xlsx.expanduser().resolve()
    output_dir: Path = args.output_dir.expanduser().resolve()

    if not input_path.is_file():
        raise SystemExit(f"入力が見つかりません: {input_path}")
    if not xlsx_path.is_file():
        raise SystemExit(f"国分類 Excel が見つかりません: {xlsx_path}")

    region_codes = load_region_codes(xlsx_path)

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    buckets = split_features(data, region_codes)

    total_written = 0
    for region in REGION_ORDER:
        feats = buckets[region]
        out_path = output_dir / f"{region}.geojson"
        write_collection(feats, out_path)
        print(f"{len(feats):6d}  features → {out_path.relative_to(root)}")
        total_written += len(feats)

    nc = buckets[NO_COUNTRY_FILE]
    nc_path = output_dir / f"{NO_COUNTRY_FILE}.geojson"
    write_collection(nc, nc_path)
    print(f"{len(nc):6d}  features → {nc_path.relative_to(root)}")
    total_written += len(nc)

    n_in = len(data.get("features") or [])
    print(f"完了: 入力 {n_in} features / 出力合計 {total_written} features")


if __name__ == "__main__":
    main()
